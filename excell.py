from deep_translator import GoogleTranslator
from dataloader import DataLoader
import openpyxl, json, time, os

DEBUG = False

sheet_folder = "Sheets"
FILES = {
	"events": os.path.join(sheet_folder, "events.xlsx"),
	"story": os.path.join(sheet_folder, "campaigns.xlsx"),
	"raids": os.path.join(sheet_folder, "raids.xlsx")
}

class Translate():

	def __init__(self):
		self.translator = GoogleTranslator(source='auto', target='en')

	def translate(self, message):
		# add a translate flag
		try:
			result = self.translator.translate(message)
			return result
		except:
			print("\n[ERROR  ]: Translation error, waiting 30 secs...\n")
			time.sleep(30)
			return self.translate(message)

class Excell():

	no_replace = False

	name_jpc = 2
	name_enc = 5
	message_jpc = 3 # japan
	message_gtc = 4 # google
	message_mnc = 6 # manual

	story_id = 26

	step = 1
	row_pos = 2
	story_number = 1
	story_begin = 0

	def __init__(self, filename):
		self.translator = Translate()
		self.filename = filename
		self.check_file(filename)
		self.file = openpyxl.load_workbook(filename)
		self.sheets = self.file.sheetnames

	def check_file(self, filename):
		if not os.path.exists(sheet_folder):
			os.makedirs(sheet_folder)
		if not os.path.exists(filename):
			file = openpyxl.Workbook()
			file.save(filename)
			file.close()

	def set_sheet(self, sheet_name):
		if not sheet_name in self.sheets:
			for s in self.sheets:
				switch = False
				for i in range(4):
					if sheet_name[i] != s[i]: 
						switch = True
						break
				if not switch:
					self.selected_sheet = self.file[s]
					self.selected_sheet.title = sheet_name
					self.reinit_sheet()
					print(f"[INFO    ]: Renaming {sheet_name} event")
					return False

			self.file.create_sheet(title=sheet_name)
			self.selected_sheet = self.file[sheet_name]# get_sheet_by_name(sheet_name)
			self.reinit_sheet()
			print(f"[INFO    ]: Add {sheet_name} event")
			return True
		else:
			self.selected_sheet = self.file[sheet_name]
			self.selected_sheet.title = sheet_name
			self.reinit_sheet()
			# print(self.selected_sheet.title)
			return False

	def set_value(self, row:str, col:int, value:str):
		self.selected_sheet[row + str(col)].value = value

	def	set_value(self, row:int, col:int, value:str):
		self.selected_sheet.cell(row=row, column=col, value=value)

	def	get_value(self, row:str, col:int):
		value = self.selected_sheet[row + str(col)].value
		# exception caught
		return value

	def get_value(self, row:int, col:int):
		return self.selected_sheet.cell(row=row, column=col)

	def save(self):
		self.file.save(self.filename)

	def	reinit_sheet(self):
		self.row_pos = 3
		self.story_number = 1
		self.set_value(1, self.name_jpc, "Name (JP)")
		self.set_value(1, self.name_enc, "Name (Manual)")
		self.set_value(1, self.message_jpc, "JP")
		self.set_value(1, self.message_gtc, "Google translate")
		self.set_value(1, self.message_mnc, "Manual translation")
		self.set_value(1, self.story_id, "Dialogue ID")

	def write_episode(self, episode_name):
		col = self.message_jpc
		row = self.row_pos

		self.set_value(row, 1, "Episode title:")
		self.set_value(row, col, episode_name)
		self.row_pos += 1

	def write_english_area(self, area, area_jp):
		col = self.message_jpc
		row = self.row_pos

		print(f"	[INFO    ]: Add {area_jp['name']} area")
		if not self.no_replace:
			self.set_value(row, 1, "Area name:")
			self.set_value(row, col, area_jp['name'])
			self.set_value(row, self.message_gtc, area['name'])
		self.set_value(row, self.story_id, area['id'])
		self.row_pos += 2
		self.story_number = 1

	def write_area(self, area):
		col = self.message_jpc
		row = self.row_pos

		print(f"	[INFO    ]: Add {area['name']} area")
		if not self.no_replace:
			self.set_value(row, 1, "Area name:")
			self.set_value(row, col, area['name'])
			self.set_value(row, self.message_gtc, self.translate_sentence(area['name']))
		self.set_value(row, self.story_id, area['id'])
		self.row_pos += 2
		self.story_number = 1

	def write_english_story(self, story, story_jp, st_name=None):
		col = self.message_jpc
		row = self.row_pos

		if not self.no_replace:
			self.set_value(row, 1, f"Scene {self.story_number}:")
			self.set_value(row, col, story_jp['title'])
			self.set_value(row, self.message_gtc, story['title'])

		self.set_value(row, self.story_id, story['id'])

		self.row_pos += 1
		self.story_begin = 0

		self.story_number += 1

	def write_story(self, story, st_name=None):
		col = self.message_jpc
		row = self.row_pos

		if not self.no_replace:
			self.set_value(row, 1, f"Scene {self.story_number}:")
			self.set_value(row, col, story['title'])
			self.set_value(row, self.message_gtc, self.translate_sentence(story['title']))

		self.set_value(row, self.story_id, story['id'])

		self.row_pos += 1
		self.story_begin = 0

		self.story_number += 1

	def	write_name(self, name):
		col = self.name_jpc
		row = self.row_pos

		if self.story_begin == 0:
			self.set_value(row, 1, "text:")
			self.story_begin = 1
		self.set_value(row, col, name)

	def write_message(self, message):
		col = self.message_jpc
		row = self.row_pos

		self.set_value(row, col, message)

	def write_story_id(self, sid):
		col = self.story_id
		row = self.row_pos

		self.set_value(row, col, sid)

	def write_english_message(self, message):
		col = self.message_gtc
		row = self.row_pos

		if not self.no_replace:
			self.set_value(row, col, message)

	def write_translated_message(self, message):
		col = self.message_gtc
		row = self.row_pos

		if not self.no_replace:
			self.set_value(row, col, self.translate_sentence(message))

		# self.row_pos += self.step

	def translate_sentence(self, message):
		return self.translator.translate(message)

	def stepping(self):
		self.row_pos += self.step

class DialogueExtractor(object):

	def __init__(self, file='japan.xlsx', jp=True, tp='event'):
		self.db = DataLoader(jp=jp)
		self.ex = Excell(file)
		
		if tp != 'raid': self.start_cycle(tp)
		else: self.raids_start_cycle()

	def	get_name(self, t):
		options = ['chara1_name', 'chara2_name', 'chara3_name']
		for o in options:
			if t[o] != "":
				return t[o]

	def is_episode(self, ep:dict) -> bool:
		return True if ep.get('episode', False) else False

	def	story_talk_cycle(self, story):
		for talk in self.db.story_talk:
			to_translate = True
			if talk['m_story_id'] != story['id']: continue
			self.ex.write_name(self.get_name(talk))
			self.ex.write_message(talk['talk_text'].replace("\n", " "))
			for talk_en in self.db.story_talk_en:
				if not talk['id'] == talk_en['id']: continue
				# print(talk_en['talk_text'])
				self.ex.write_english_message(talk_en['talk_text'].replace("\n", " "))
				to_translate = False
				break
			if to_translate:
				self.ex.write_translated_message(talk['talk_text'].replace("\n", " "))
			self.ex.write_story_id(talk['id'])
			self.ex.row_pos += self.ex.step

	def	story_cycle(self, area, ep):

		def write_story(story):
			for story_en in self.db.story_en:
				if not story_en['id'] == story['id']: continue
				self.ex.write_english_story(story_en, story)
				return
			self.ex.write_story(story)

		for story in self.db.story:
			divider = 100 if ep['event_type'] == 1 else 10
			if self.is_episode(ep):
				if (int(story['id'] / 100) != area['m_episode_id']): continue
				if str(story['id'])[-2] != str(area['id'])[-1]: continue
			else:
				if int(story['id'] / divider) != area['id']: continue
			write_story(story)
			self.story_talk_cycle(story)
			self.ex.row_pos += 1

	def	area_cycle(self, ep):

		def write_area(area):
			for area_en in self.db.area_en:
				if not area_en['id'] == area['id']: continue
				self.ex.write_english_area(area_en, area)
				return
			self.ex.write_area(area)

		area_num = 0
		for area in self.db.area:
			if area['m_episode_id'] != ep['m_episode_id']: continue
			area_num += 1
			area_name = area['name']

			write_area(area)
			self.story_cycle(area, ep)

	def start_cycle(self, tp='event'):
		event_types = [1, 15]
		story = self.db.event if tp == 'event' else self.db.episode 
		for ep in story:
			ep['episode'] = False
			if not ep[f'{tp}_type'] in event_types: continue
			if not 'm_episode_id' in ep:
				ep['m_episode_id'] = ep['id']
				ep['event_type'] = ep['episode_type']
				ep['episode'] = True

			ename = ep['resource_name'] if tp == 'event' else ep['name']
			sheet_name = f"{str(ep['id']).rjust(3, '0')}. {ename}"
			if not self.ex.set_sheet(sheet_name) and not DEBUG: continue
			self.area_cycle(ep)
			self.ex.save()

		self.ex.save()

	def raid_story_cycle(self, ep) -> None:
		for story in self.db.story:
			if story['id'] != ep['prologue_story_id'] and story['id'] != ep['ending_story_id']: continue
			pre = 'prologue' if story['id'] == ep['prologue_story_id'] else 'ending'
			story['title'] = f"{pre}: {story['title']}"

			self.ex.write_story(story)
			self.story_talk_cycle(story)
			self.ex.row_pos += 1

	def raids_start_cycle(self):
		done_raids = []
		story = self.db.event
		for ep in story:
			if not ep['event_type'] == 6: continue
			if ((ep['resource_name'][-1] == 'e') or
				(ep['resource_name'][-1] == 'c') or
				(ep['resource_name'] in done_raids)):
				continue

			done_raids.append(ep['resource_name'])
			sheet_name = f"{str(ep['id']).rjust(3, '0')}. {ep['resource_name'].replace(' ', '_')}"
			if not self.ex.set_sheet(sheet_name) and not DEBUG: continue

			self.raid_story_cycle(ep)
			self.ex.save()
		self.ex.save()

# replace occurences in json
class DialogueReverse():

	def __init__(self):
		pass


if __name__ == '__main__':
	# Extract Campaign Story
	DialogueExtractor(file=FILES['story'], jp=True, tp='episode')
	# Extract Story Events
	DialogueExtractor(file=FILES['events'], jp=True, tp='event')
	# Extract Raids prologue-ending
	DialogueExtractor(file=FILES['raids'], jp=True, tp='raid')